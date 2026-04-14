from __future__ import annotations

import argparse
import re
from dataclasses import dataclass, field
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
DEFAULT_DOC_ROOT = Path(r"D:\CoWorkProjects\需求分析文档")
DEFAULT_OUTPUT_DIR = Path("assessment_bank_exports")
SURVEY_1_NAME = "第一次问卷采集表.docx"
SURVEY_2_NAME = "第二次问卷采集表.docx"
EXAM_NAME = "科研考试和思政考试试卷.docx"
OPTION_LABELS = [chr(code) for code in range(ord("A"), ord("Z") + 1)]
OPTION_HEADERS = [f"option_{label.lower()}" for label in OPTION_LABELS]


@dataclass
class ExportQuestion:
    question_no: int
    question_type: str
    title: str
    options: list[str] = field(default_factory=list)
    answer: str = ""
    score: int = 0


@dataclass
class ExportPaper:
    paper_type: str
    title: str
    version_no: int
    duration_seconds: int
    source_file: str
    questions: list[ExportQuestion] = field(default_factory=list)


def read_docx_lines(path: Path) -> list[str]:
    with ZipFile(path) as archive:
        xml_text = archive.read("word/document.xml")

    root = ET.fromstring(xml_text)
    lines: list[str] = []
    for paragraph in root.findall(".//w:p", WORD_NS):
        text = "".join(node.text or "" for node in paragraph.findall(".//w:t", WORD_NS)).strip()
        if text:
            lines.append(normalize_line(text))
    return lines


def normalize_line(line: str) -> str:
    replacements = {
        "\u3000": " ",
        "\u00a0": " ",
        "“": '"',
        "”": '"',
        "‘": "'",
        "’": "'",
        "•": "",
        "": "",
        "□": "■",
    }
    normalized = line
    for source, target in replacements.items():
        normalized = normalized.replace(source, target)
    return re.sub(r"\s+", " ", normalized).strip()


def normalize_question_prompt(text: str) -> str:
    return re.sub(r"^\d+[.．、\s]*", "", text).strip().rstrip("：:")


def extract_checkbox_options(text: str) -> tuple[str, list[str]]:
    if "■" not in text:
        return text.strip(), []

    first_box = text.find("■")
    prompt = text[:first_box].strip().rstrip("（）:：")
    option_part = text[first_box:]
    options = [
        item.strip()
        for item in re.findall(r"■\s*([^■]+?)(?=\s*■|$)", option_part)
        if item.strip()
    ]
    return prompt, options


def build_option_cells(options: list[str]) -> list[str]:
    padded = (options + [""] * len(OPTION_HEADERS))[: len(OPTION_HEADERS)]
    return padded


def survey_heading_pattern() -> re.Pattern[str]:
    return re.compile(r"^[一二三四五六七八九十]+[、.．]")


def is_survey_heading(line: str) -> bool:
    if survey_heading_pattern().match(line):
        return True
    return line in {
        "高中",
        "本科",
        "硕士研究生（无则填 “无”）",
        "硕士研究生（无则填“无”）",
        "博士研究生（无则填 “无”）",
        "博士研究生（无则填“无”）",
    }


def infer_survey_question_type(prompt: str, options: list[str]) -> str:
    if not options:
        return "fill_blank"
    if any(keyword in prompt for keyword in ("可多选", "限选", "多选", "限选 3 项", "限选3项")):
        return "multiple"
    return "single"


def parse_survey_doc(path: Path, title: str, version_no: int) -> ExportPaper:
    paper = ExportPaper(
        paper_type="survey",
        title=title,
        version_no=version_no,
        duration_seconds=5400,
        source_file=path.name,
    )
    question_no = 1
    pending_prompt: str | None = None
    pending_options: list[str] = []

    def flush_pending() -> None:
        nonlocal pending_prompt, pending_options, question_no
        if not pending_prompt:
            pending_options = []
            return
        question_type = infer_survey_question_type(pending_prompt, pending_options)
        paper.questions.append(
            ExportQuestion(
                question_no=question_no,
                question_type=question_type,
                title=pending_prompt,
                options=list(pending_options),
            )
        )
        question_no += 1
        pending_prompt = None
        pending_options = []

    for line in read_docx_lines(path):
        if not line or is_survey_heading(line):
            flush_pending()
            continue

        prompt, options = extract_checkbox_options(line)
        has_number_prefix = bool(re.match(r"^\d+[.．、\s]*", line))
        is_fill_blank = "____" in line or "________" in line
        is_prompt_like = line.endswith(("：", ":"))

        if has_number_prefix:
            flush_pending()
            if options:
                pending_prompt = normalize_question_prompt(prompt)
                pending_options = list(options)
            elif is_fill_blank:
                paper.questions.append(
                    ExportQuestion(
                        question_no=question_no,
                        question_type="fill_blank",
                        title=normalize_question_prompt(line).rstrip("（）"),
                    )
                )
                question_no += 1
            else:
                pending_prompt = normalize_question_prompt(line)
                pending_options = []
            continue

        if options:
            if prompt:
                flush_pending()
                pending_prompt = normalize_question_prompt(prompt)
                pending_options = list(options)
            elif pending_prompt:
                pending_options.extend(options)
            continue

        if is_fill_blank and not options:
            flush_pending()
            paper.questions.append(
                ExportQuestion(
                    question_no=question_no,
                    question_type="fill_blank",
                    title=normalize_question_prompt(line).rstrip("（）"),
                )
            )
            question_no += 1
            continue

        if is_prompt_like:
            flush_pending()
            pending_prompt = normalize_question_prompt(line)
            pending_options = []
            continue

        if pending_prompt:
            pending_prompt = f"{pending_prompt} {line}".strip()

    flush_pending()
    return paper


def parse_exam_heading(line: str) -> tuple[str, int] | None:
    patterns = [
        (r"^科研诚信试卷（第\s*(\d+)\s*套）$", "integrity"),
        (r"^思政试卷（第\s*(\d+)\s*套）$", "ideology"),
    ]
    for pattern, paper_type in patterns:
        match = re.match(pattern, line)
        if match:
            return paper_type, int(match.group(1))
    return None


def detect_exam_question_type(line: str) -> str | None:
    if line.startswith(("一、单选题", "单选题")):
        return "single"
    if line.startswith(("二、多选题", "多选题")):
        return "multiple"
    if line.startswith(("三、判断题", "判断题")):
        return "boolean"
    return None


def is_exam_question_line(line: str) -> bool:
    return bool(re.match(r"^\d+[.．、\s]", line))


def is_exam_prompt_line(line: str) -> bool:
    if line.startswith(tuple(f"{label}." for label in OPTION_LABELS)):
        return False
    return line.endswith("（）") or line.endswith("()")


def clean_exam_question_title(line: str) -> str:
    title = re.sub(r"^\d+[.．、\s]*", "", line).strip()
    title = re.sub(r"[（(]\s*[）)]\s*$", "", title).strip()
    return title


def split_exam_options(text: str) -> list[str]:
    matches = list(re.finditer(r"([A-Z])\.\s*", text))
    if not matches:
        return []

    values: list[str] = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        values.append(text[start:end].strip(" ；;"))
    return values


def parse_exam_doc(path: Path) -> list[ExportPaper]:
    papers: list[ExportPaper] = []
    current_paper: ExportPaper | None = None
    current_question_type: str | None = None
    current_question: ExportQuestion | None = None
    current_question_no = 1

    def flush_question() -> None:
        nonlocal current_question
        if current_paper and current_question:
            current_paper.questions.append(current_question)
        current_question = None

    for raw_line in read_docx_lines(path):
        line = raw_line.strip()
        if not line:
            continue

        if "答案 + 解析" in line or "答案+解析" in line:
            flush_question()
            break

        heading = parse_exam_heading(line)
        if heading:
            flush_question()
            paper_type, version_no = heading
            title_prefix = "科研诚信试卷" if paper_type == "integrity" else "思政试卷"
            current_paper = ExportPaper(
                paper_type=paper_type,
                title=f"{title_prefix}（第{version_no}套）",
                version_no=version_no,
                duration_seconds=7200,
                source_file=path.name,
            )
            papers.append(current_paper)
            current_question_type = None
            current_question_no = 1
            continue

        detected_question_type = detect_exam_question_type(line)
        if detected_question_type:
            flush_question()
            current_question_type = detected_question_type
            continue

        if not current_paper or not current_question_type:
            continue

        if current_question_type in {"single", "multiple"} and re.search(r"\bA\.\s*", line):
            if current_question:
                current_question.options = split_exam_options(line)
            continue

        if is_exam_question_line(line) or is_exam_prompt_line(line):
            flush_question()
            current_question = ExportQuestion(
                question_no=current_question_no,
                question_type=current_question_type,
                title=clean_exam_question_title(line),
                options=["对", "错"] if current_question_type == "boolean" else [],
                score=2,
            )
            current_question_no += 1
            continue

        if current_question:
            current_question.title = f"{current_question.title} {line}".strip()

    flush_question()
    return papers


def build_default_papers(doc_root: Path) -> list[ExportPaper]:
    papers = [
        parse_survey_doc(doc_root / SURVEY_1_NAME, "德育画像构建问卷一", 1),
        parse_survey_doc(doc_root / SURVEY_2_NAME, "德育画像构建问卷二", 2),
    ]
    papers.extend(parse_exam_doc(doc_root / EXAM_NAME))
    return papers


def safe_filename(value: str) -> str:
    value = value.strip().replace(" ", "_")
    value = re.sub(r"[\\/:*?\"<>|]", "_", value)
    return value


def build_output_filename(paper: ExportPaper) -> str:
    prefix_map = {
        "survey": "问卷",
        "integrity": "科研诚信试卷",
        "ideology": "思政试卷",
    }
    prefix = prefix_map.get(paper.paper_type, paper.paper_type)
    return safe_filename(f"{prefix}_第{paper.version_no}份.xlsx")


def autofit_sheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[column].width = min(max(length + 2, 12), 40)


def write_single_workbook(output_path: Path, paper: ExportPaper) -> None:
    workbook = Workbook()

    readme = workbook.active
    readme.title = "README"
    readme.append(["说明"])
    readme.append(["本文件由 docx 题库自动解析生成，请先人工检查题干、题型、选项、答案和分值。"])
    readme.append(["如有问题，可直接在 questions 工作表里手动修改。"])
    readme.append([f"paper_type={paper.paper_type}, title={paper.title}, version_no={paper.version_no}"])

    paper_sheet = workbook.create_sheet("paper")
    paper_sheet.append(["paper_type", "paper_title", "version_no", "duration_seconds", "question_count", "source_file"])
    paper_sheet.append(
        [
            paper.paper_type,
            paper.title,
            paper.version_no,
            paper.duration_seconds,
            len(paper.questions),
            paper.source_file,
        ]
    )

    questions_sheet = workbook.create_sheet("questions")
    questions_sheet.append(
        [
            "paper_type",
            "paper_title",
            "version_no",
            "question_no",
            "question_type",
            "question_title",
            *OPTION_HEADERS,
            "answer",
            "score",
            "source_file",
        ]
    )
    for question in paper.questions:
        questions_sheet.append(
            [
                paper.paper_type,
                paper.title,
                paper.version_no,
                question.question_no,
                question.question_type,
                question.title,
                *build_option_cells(question.options),
                question.answer,
                question.score,
                paper.source_file,
            ]
        )

    header_font = Font(bold=True)
    for worksheet in (readme, paper_sheet, questions_sheet):
        for cell in worksheet[1]:
            cell.font = header_font
        worksheet.freeze_panes = "A2"
        autofit_sheet(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def export_workbooks(output_dir: Path, papers: list[ExportPaper]) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    exported_paths: list[Path] = []
    for paper in papers:
        path = output_dir / build_output_filename(paper)
        write_single_workbook(path, paper)
        exported_paths.append(path)
    return exported_paths


def print_summary(exported_paths: list[Path], papers: list[ExportPaper]) -> None:
    print(f"Exported files: {len(exported_paths)}")
    for path, paper in zip(exported_paths, papers):
        max_options = max((len(question.options) for question in paper.questions), default=0)
        print(
            f"- {path.resolve()} | {paper.paper_type} v{paper.version_no} | "
            f"{len(paper.questions)} questions | max_options={max_options}"
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export questionnaire and exam content from DOCX files to Excel.")
    parser.add_argument(
        "--doc-root",
        type=Path,
        default=DEFAULT_DOC_ROOT,
        help="Folder containing the provided DOCX source files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Output directory for generated .xlsx files.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    papers = build_default_papers(args.doc_root)
    exported_paths = export_workbooks(args.output_dir, papers)
    print_summary(exported_paths, papers)


if __name__ == "__main__":
    main()
