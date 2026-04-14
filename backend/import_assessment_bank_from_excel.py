from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.database import Base, SessionLocal, engine
from app.models.assessment_paper import AssessmentPaper
from app.models.assessment_paper_question import AssessmentPaperQuestion
from app.models.assessment_question import AssessmentQuestion
from app.models.auth_user import AuthUser  # noqa: F401
from app.models.student_user import StudentUser  # noqa: F401
from app.models.teacher_user import TeacherUser  # noqa: F401


DEFAULT_INPUT_DIR = Path("assessment_bank_exports")
SUPPORTED_PAPER_TYPES = {"survey", "integrity", "ideology"}
SUPPORTED_QUESTION_TYPES = {"single", "multiple", "boolean", "fill_blank", "essay"}


@dataclass
class ImportedQuestionRow:
    question_no: int
    question_type: str
    title: str
    option_a: str = ""
    option_b: str = ""
    option_c: str = ""
    option_d: str = ""
    answer: str = ""
    score: int = 0


@dataclass
class ImportedPaperSheet:
    paper_type: str
    title: str
    version_no: int
    duration_seconds: int
    source_file: str
    questions: list[ImportedQuestionRow]


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_options_json(row: ImportedQuestionRow) -> list[dict[str, str]] | None:
    options = []
    for label, text in (
        ("A", row.option_a),
        ("B", row.option_b),
        ("C", row.option_c),
        ("D", row.option_d),
    ):
        if text:
            options.append({"label": label, "text": text})
    return options or None


def build_answer_json(row: ImportedQuestionRow) -> Any:
    answer = normalize_cell(row.answer)
    if not answer:
        return None

    if row.question_type == "multiple":
        return [item.strip() for item in answer.split(",") if item.strip()]
    return answer


def parse_excel_file(path: Path) -> ImportedPaperSheet:
    workbook = load_workbook(path)
    if "paper" not in workbook.sheetnames or "questions" not in workbook.sheetnames:
        raise ValueError(f"{path.name} 缺少必要的工作表: paper/questions")

    paper_sheet = workbook["paper"]
    questions_sheet = workbook["questions"]

    paper_headers = [normalize_cell(cell.value) for cell in next(paper_sheet.iter_rows(min_row=1, max_row=1))]
    paper_values = [cell.value for cell in next(paper_sheet.iter_rows(min_row=2, max_row=2))]
    paper_data = dict(zip(paper_headers, paper_values))

    paper_type = normalize_cell(paper_data.get("paper_type"))
    title = normalize_cell(paper_data.get("paper_title"))
    version_no = int(paper_data.get("version_no") or 0)
    duration_seconds = int(paper_data.get("duration_seconds") or 0)
    source_file = normalize_cell(paper_data.get("source_file"))

    if paper_type not in SUPPORTED_PAPER_TYPES:
        raise ValueError(f"{path.name} 的 paper_type 不支持: {paper_type}")
    if not title:
        raise ValueError(f"{path.name} 缺少 paper_title")
    if version_no <= 0:
        raise ValueError(f"{path.name} 的 version_no 非法: {version_no}")

    question_headers = [normalize_cell(cell.value) for cell in next(questions_sheet.iter_rows(min_row=1, max_row=1))]
    questions: list[ImportedQuestionRow] = []
    for row in questions_sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(question_headers, row))
        question_type = normalize_cell(row_data.get("question_type"))
        title = normalize_cell(row_data.get("question_title"))
        if not question_type and not title:
            continue
        if question_type not in SUPPORTED_QUESTION_TYPES:
            raise ValueError(f"{path.name} 存在不支持的 question_type: {question_type}")
        if not title:
            raise ValueError(f"{path.name} 存在空 question_title")

        questions.append(
            ImportedQuestionRow(
                question_no=int(row_data.get("question_no") or 0),
                question_type=question_type,
                title=title,
                option_a=normalize_cell(row_data.get("option_a")),
                option_b=normalize_cell(row_data.get("option_b")),
                option_c=normalize_cell(row_data.get("option_c")),
                option_d=normalize_cell(row_data.get("option_d")),
                answer=normalize_cell(row_data.get("answer")),
                score=int(row_data.get("score") or 0),
            )
        )

    questions.sort(key=lambda item: item.question_no)
    return ImportedPaperSheet(
        paper_type=paper_type,
        title=normalize_cell(paper_data.get("paper_title")),
        version_no=version_no,
        duration_seconds=duration_seconds,
        source_file=source_file or path.name,
        questions=questions,
    )


def upsert_paper(
    db: Session,
    *,
    paper_type: str,
    title: str,
    version_no: int,
    duration_seconds: int,
) -> AssessmentPaper:
    paper = (
        db.query(AssessmentPaper)
        .filter(
            AssessmentPaper.paper_type == paper_type,
            AssessmentPaper.version_no == version_no,
        )
        .first()
    )

    if paper:
        paper.title = title
        paper.duration_seconds = duration_seconds
        paper.is_active = True
        return paper

    paper = AssessmentPaper(
        paper_type=paper_type,
        title=title,
        version_no=version_no,
        duration_seconds=duration_seconds,
        is_active=True,
        created_by_user_id=None,
    )
    db.add(paper)
    db.flush()
    return paper


def clear_existing_questions(db: Session, paper: AssessmentPaper) -> None:
    question_ids = [
        item.question_id
        for item in db.query(AssessmentPaperQuestion.question_id)
        .filter(AssessmentPaperQuestion.paper_id == paper.id)
        .all()
    ]
    db.query(AssessmentPaperQuestion).filter(AssessmentPaperQuestion.paper_id == paper.id).delete()
    db.flush()
    if question_ids:
        db.query(AssessmentQuestion).filter(AssessmentQuestion.id.in_(question_ids)).delete(synchronize_session=False)
        db.flush()


def create_question(db: Session, paper: AssessmentPaper, row: ImportedQuestionRow) -> AssessmentQuestion:
    question = AssessmentQuestion(
        paper_type=paper.paper_type,
        question_type=row.question_type,
        title=row.title,
        options_json=build_options_json(row),
        answer_json=build_answer_json(row),
        score=row.score,
        created_by_user_id=None,
    )
    db.add(question)
    db.flush()
    return question


def link_question(db: Session, paper: AssessmentPaper, question: AssessmentQuestion, sort_order: int) -> None:
    db.add(
        AssessmentPaperQuestion(
            paper_id=paper.id,
            question_id=question.id,
            sort_order=sort_order,
        )
    )


def import_excel_file(db: Session, path: Path) -> tuple[str, int]:
    paper_sheet = parse_excel_file(path)
    paper = upsert_paper(
        db,
        paper_type=paper_sheet.paper_type,
        title=paper_sheet.title,
        version_no=paper_sheet.version_no,
        duration_seconds=paper_sheet.duration_seconds,
    )
    clear_existing_questions(db, paper)

    for index, row in enumerate(paper_sheet.questions, start=1):
        question = create_question(db, paper, row)
        link_question(db, paper, question, index)

    db.flush()
    return paper.title, len(paper_sheet.questions)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import assessment bank Excel files into database.")
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help="Directory containing reviewed assessment bank .xlsx files.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate Excel files without writing to database.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    excel_files = sorted(path for path in args.input_dir.glob("*.xlsx") if not path.name.startswith("~$"))
    if not excel_files:
        raise FileNotFoundError(f"未找到 Excel 文件: {args.input_dir}")

    if args.dry_run:
        print("Dry run complete.")
        for path in excel_files:
            paper = parse_excel_file(path)
            print(f"- {path.name} | {paper.paper_type} v{paper.version_no} | {len(paper.questions)} questions")
        return

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        imported: list[tuple[str, int, str]] = []
        for path in excel_files:
            title, question_count = import_excel_file(db, path)
            imported.append((path.name, question_count, title))
        db.commit()
        print("Import complete.")
        for file_name, question_count, title in imported:
            print(f"- {file_name} | {title} | {question_count} questions")
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
